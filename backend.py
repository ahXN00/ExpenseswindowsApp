import json
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# File to store the data
data_file = 'expenses.json'

# Initialize the file if it does not exist
if not os.path.exists(data_file):
    with open(data_file, 'w') as file:
        json.dump([], file)

def load_data():
    try:
        with open(data_file, 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        return []

def save_data(data):
    with open(data_file, 'w') as file:
        json.dump(data, file, indent=4)

def add_expense(date, category, amount):
    expenses = load_data()
    expenses.append({'date': date, 'category': category, 'amount': amount})
    save_data(expenses)

def view_monthly_expenses(month):
    expenses = pd.DataFrame(load_data())
    if not expenses.empty:
        expenses['month'] = expenses['date'].apply(lambda x: datetime.strptime(x, '%d-%m-%Y').strftime('%m-%Y'))
        monthly_data = expenses[expenses['month'] == month]
        if not monthly_data.empty:
            return monthly_data.to_string(index=False)
    return "No expenses found for this month."


def view_total_expenses():
    expenses = pd.DataFrame(load_data())
    if not expenses.empty:
        expenses['date'] = pd.to_datetime(expenses['date'], format='%d-%m-%Y')
        expenses['amount'] = pd.to_numeric(expenses['amount'], errors='coerce')
        expenses['month-year'] = expenses['date'].dt.strftime('%m-%Y')
        expenses['year'] = expenses['date'].dt.year

        # Group by year and then month within each year
        grouped = expenses.groupby(['year', 'month-year'])['amount'].sum().reset_index(name='Total')

        # Calculate differences within each year
        grouped['difference'] = grouped.groupby('year')['Total'].pct_change().fillna(0).apply(
            lambda x: f"{x * 100:.2f}% more" if x > 0 else f"{x * 100:.2f}% less")

        # Creating a summary for each year
        annual_summary = grouped.groupby('year')['Total'].sum().reset_index(name='Annual Total')
        annual_summary = annual_summary.merge(grouped, on='year', how='left')

        return annual_summary.to_string(index=False)
    return "No expenses to display."


def save_to_excel(filepath):
    data = load_data()
    if not data:
        return "No data to save."

    df = pd.DataFrame(data)
    df['date'] = pd.to_datetime(df['date'], format='%d-%m-%Y')
    df['month-year'] = df['date'].dt.strftime('%m-%Y')
    df['year'] = df['date'].dt.year
    df['amount'] = pd.to_numeric(df['amount'], errors='coerce')

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Group by year and then month within each year
        grouped = df.groupby(['year', 'month-year'])

        # Create sheets for each year and month
        for (year, month_year), group in grouped:
            sheet_name = f"{year}_{month_year}"
            group.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            adjust_column_widths(group, worksheet)
            total = group['amount'].sum()
            last_row = len(group) + 2
            worksheet.cell(row=last_row, column=1, value='Total')
            worksheet.cell(row=last_row, column=len(group.columns), value=total)

        # Create a summary sheet for each year
        summary_df = grouped['amount'].sum().reset_index(name='Total')
        summary_df['difference'] = summary_df.groupby('year')['Total'].pct_change().fillna(0).apply(
            lambda x: f"{x:.2%} more" if x > 0 else f"{x:.2%} less")

        for year, group in summary_df.groupby('year'):
            summary_sheet_name = f"Summary {year}"
            group.to_excel(writer, sheet_name=summary_sheet_name, index=False)
            summary_worksheet = writer.sheets[summary_sheet_name]
            adjust_column_widths(group, summary_worksheet)
            total_year = group['Total'].sum()
            last_row = len(group) + 2
            summary_worksheet.cell(row=last_row, column=1, value='Total for the Year')
            summary_worksheet.cell(row=last_row, column=3, value=total_year)

    return "Data successfully saved to Excel at " + filepath


def adjust_column_widths(dataframe, worksheet):
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        length = max(length, max(len(str(dataframe.columns[column_cells[0].column - 1])), 10))
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

def clear_data(option, date=None):
    data = load_data()
    if option == 'all':
        save_data([])
    elif option == 'month':
        data = [entry for entry in data if datetime.strptime(entry['date'], '%d-%m-%Y').strftime('%m-%Y') != date]
        save_data(data)
    elif option == 'date':
        data = [entry for entry in data if entry['date'] != date]
        save_data(data)


def load_from_excel(filepath):
    """Load expenses data from Excel, merge with existing data."""
    try:
        xls = pd.ExcelFile(filepath)
        new_data = []

        # Iterate through each sheet in the Excel file
        for sheet_name in xls.sheet_names:
            # Skipping summary sheets based on expected naming convention
            if 'Summary' in sheet_name:
                continue

            # Load data from the sheet
            df = pd.read_excel(xls, sheet_name=sheet_name)
            # Validate expected columns
            expected_columns = {'date', 'category', 'amount'}
            if not expected_columns.issubset(df.columns):
                continue  # Skip sheets that do not have the required columns

            # Ensure date format is correct and filter out potential parsing errors
            df['date'] = pd.to_datetime(df['date'], errors='coerce')
            df = df.dropna(subset=['date'])  # Drop rows where dates could not be parsed
            df['date'] = df['date'].dt.strftime('%d-%m-%Y')

            new_data.extend(df.to_dict('records'))

        existing_data = load_data()
        merged_data = merge_data(existing_data, new_data)
        save_data(merged_data)
        return "Data loaded and merged successfully from Excel."
    except Exception as e:
        return f"Failed to load data from Excel: {str(e)}"


def merge_data(existing_data, new_data):
    existing_df = pd.DataFrame(existing_data)
    new_df = pd.DataFrame(new_data)

    if existing_df.empty:
        return new_data
    if new_df.empty:
        return existing_data

    # Create a unique key for each entry based on date and category to avoid duplicates
    existing_df['key'] = existing_df['date'] + '_' + existing_df['category']
    new_df['key'] = new_df['date'] + '_' + new_df['category']

    # Combine and remove duplicates, keeping the latest entry
    combined_df = pd.concat([existing_df, new_df]).drop_duplicates(subset='key', keep='last')
    return combined_df.drop('key', axis=1).to_dict('records')


# Example usage
file_path = 'C:/path/to/your/excel_file.xlsx'
if os.path.exists(file_path):
    print("File found!")
    result = load_from_excel(file_path)
    print(result)
else:
    print("File not found at:", file_path)

